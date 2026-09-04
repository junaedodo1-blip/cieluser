import os, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

catalog_path = r'data/ai_visual_prompt_cookbook/catalog.json'
with open(catalog_path, 'r', encoding='utf-8') as f:
    catalog = json.load(f)

style_slugs = [item['slug'] for item in catalog]

days_data = [
    {
        'day': 1, 'day_name': 'Monday', 'pillar': 'Worldbuilding & Cultural Gravity',
        'topic': 'Your Brand is a Destination, Your Product is the Souvenir',
        'hook': 'most founders build a business. the ones who win build a world.',
        'script': 'Act I: Show why posting random promo graphics leaves your brand unremembered.\nAct II: Contrast flat product feature lists with full immersive worldbuilding.\nAct III: The product lands as the physical souvenir of your brand world.',
        'slides': 8, 'trigger': 'WORLD', 'offer': 'Brand Worldbuilding Blueprint PDF & Figma Template'
    },
    {
        'day': 2, 'day_name': 'Tuesday', 'pillar': 'Auteur Brand Transformation Arc',
        'topic': 'Nobody Remembers Your Origin Story: The Psychology of Transformation',
        'hook': 'nobody cares about your founding story. they care about who they become.',
        'script': 'Act I: Unspoken truth—customers tune out resume origins.\nAct II: Deconstruct the human tension of wanting an elevated identity.\nAct III: Position your product as the catalyst for their identity shift.',
        'slides': 7, 'trigger': 'TRANSFORM', 'offer': '3-Act Brand Narrative Playbook & Prompt Pack'
    },
    {
        'day': 3, 'day_name': 'Wednesday', 'pillar': 'Visual Directing & Cinema Optics',
        'topic': 'Why Deep Shadow Wells & Negative Space Create Subconscious Luxury Value',
        'hook': 'cheap visuals flood the frame with light. auteur cinema uses shadow to create luxury.',
        'script': 'Act I: Flat studio lighting destroys perceived product value.\nAct II: Explain raking side key light, 50mm optical prime compression, and deep shadow wells.\nAct III: Copy-paste prompt formula for 4K luxury lighting.',
        'slides': 6, 'trigger': 'CINEMA', 'offer': '50mm Cinema Lighting Code Pack'
    },
    {
        'day': 4, 'day_name': 'Thursday', 'pillar': 'Uncanny Luxury Juxtaposition',
        'topic': 'High-Fashion Tailoring in Brutalist Concrete: The Power of Unexpected Contrast',
        'hook': 'why placing luxury in raw concrete generates 10x more engagement than a studio.',
        'script': 'Act I: Clean studio shots feel generic and commercial.\nAct II: Deconstruct the Jacquemus/Balenciaga rule of high-friction contrast.\nAct III: 4 Juxtaposition Archetypes with copy-paste prompts.',
        'slides': 8, 'trigger': 'JUXTA', 'offer': 'Uncanny Juxtaposition Prompt Matrix'
    },
    {
        'day': 5, 'day_name': 'Friday', 'pillar': 'The Invisible Craft & Auteur Moat',
        'topic': 'Why We Spent 3 Months Obsessed Over a 0.5-Second Animation',
        'hook': 'spectacle without soul is forgotten in three seconds.',
        'script': 'Act I: AI makes spectacle easy, but soul remains rare.\nAct II: Walk through the friction of spending months tuning spring physics & macro light caustics.\nAct III: Why obsession over invisible details creates an untouchable moat.',
        'slides': 8, 'trigger': 'CRAFT', 'offer': 'Soul Cinema Directing Playbook'
    },
    {
        'day': 6, 'day_name': 'Saturday', 'pillar': 'Niche Specificity & Obsession',
        'topic': 'Be SO Specific It Feels Risky: The Maison Balzac Case Study',
        'hook': 'none of the obsession-worthy brands set out to be loved by everyone.',
        'script': 'Act I: Broad appealing brands become options; specific brands become magnets.\nAct II: Deconstruct how 5 memory-inspired candle scents built an iconic cult brand.\nAct III: Step-by-step niche positioning framework.',
        'slides': 9, 'trigger': 'MAGNETIC', 'offer': 'Obsession Brand Positioning Guide'
    },
    {
        'day': 7, 'day_name': 'Sunday', 'pillar': 'Visionary Founder Arc',
        'topic': 'Why Great Founders Never Pitch What They Built: The Steve Jobs Formula',
        'hook': 'your pitch is losing everyone in minute two.',
        'script': 'Act I: Listing tech specs and feature modules causes audience fatigue.\nAct II: Reframe 5GB storage into 1,000 songs in your pocket.\nAct III: Founder narrative architecture cheat sheet.',
        'slides': 7, 'trigger': 'FOUNDER', 'offer': 'Visionary Founder Pitch Architecture'
    },
    {
        'day': 8, 'day_name': 'Monday', 'pillar': 'Sensory Physics & Viscosity',
        'topic': 'Macro Liquid Viscosity & 1/10,000s Shutter Physics in AI Cinema',
        'hook': 'why normal ai product videos look fake—and how liquid viscosity fixes them.',
        'script': 'Act I: Constant linear speed reveals synthetic CGI.\nAct II: Spring math, liquid droplet freeze, and deep sub-bass impact audio.\nAct III: Complete Seedance/Higgsfield video prompt code.',
        'slides': 6, 'trigger': 'PHYSICS', 'offer': 'Macro Fluid Physics Prompt Code'
    },
    {
        'day': 9, 'day_name': 'Tuesday', 'pillar': 'Contrarian Agency Strategy',
        'topic': 'Why 90% of Visual Campaigns Fail: The Blank Prompt Fallacy',
        'hook': 'typing random buzzwords gives you chaotic, messy ads.',
        'script': 'Act I: The mistake of starting without a strict visual blueprint.\nAct II: How elite design agencies lock grid structure, font weights, and color balance first.\nAct III: The 5-Layer Structural Prompt Specification.',
        'slides': 8, 'trigger': 'GRID', 'offer': '5-Layer Layout Specification Sheet'
    },
    {
        'day': 10, 'day_name': 'Wednesday', 'pillar': 'Automated DM Conversion Engine',
        'topic': 'How to Turn 1 Instagram Comment into ,000 in Automated Sales 24/7',
        'hook': 'stop putting links in your bio. here is what top personal brands do instead.',
        'script': 'Act I: Bio links have 95% dropoff rate.\nAct II: The 2-second OpenReply comment-to-DM automated lead magnet funnel.\nAct III: 1-click copy-paste automation script.',
        'slides': 8, 'trigger': 'FUNNEL', 'offer': 'OpenReply 24/7 Automated Funnel Guide'
    },
    {
        'day': 11, 'day_name': 'Thursday', 'pillar': 'Swiss HUD Data Teardown',
        'topic': 'Clinical Biotech & 0.5px Data HUD Diagrams for Physical Products',
        'hook': 'how 0.5px swiss hud metadata makes physical objects look clinical & scientific.',
        'script': 'Act I: Fluff marketing claims fail in 2026.\nAct II: Exploded material anatomy, GSM, scent pyramids, and technical data HUD framing.\nAct III: Swiss HUD Figma template & prompt tags.',
        'slides': 7, 'trigger': 'BIOTECH', 'offer': 'Swiss HUD Metadata Figma Kit'
    },
    {
        'day': 12, 'day_name': 'Friday', 'pillar': 'Urban Street POV & Hard Truths',
        'topic': 'Dark Mode Apple Notes & Hard Truths: The Raw POV Formula',
        'hook': 'why raw street pov notes get 5x higher save rates than polished corporate slides.',
        'script': 'Act I: Corporate PR polish creates skepticism.\nAct II: Pixelated face blur + raw Apple Notes hot take on industry lies.\nAct III: The Raw POV engagement formula.',
        'slides': 5, 'trigger': 'RAWPOV', 'offer': 'Apple Notes Editorial Template'
    },
    {
        'day': 13, 'day_name': 'Saturday', 'pillar': 'Scale Disruption & Surreal Art',
        'topic': 'Giant 3D Inflatable Foil Balloons in Underground Elevators: The Jacquemus Playbook',
        'hook': 'scale disruption: how giant objects in unexpected places hijack human attention.',
        'script': 'Act I: Standard size products get scrolled past.\nAct II: Explaining scale manipulation, 3D inflatable textures, and ambient lighting.\nAct III: Copy-paste inflatable 3D prompt formula.',
        'slides': 8, 'trigger': 'BALLOON', 'offer': 'Scale Disruption Prompt Pack'
    },
    {
        'day': 14, 'day_name': 'Sunday', 'pillar': 'Skate & Streetwear Decal Drops',
        'topic': 'Seamless Studio Hero Objects Plastered in Skate & Streetwear Decal Stickers',
        'hook': 'how physical streetwear decals turn luxury objects into cult items.',
        'script': 'Act I: Minimalist products can feel sterile without human culture.\nAct II: Layering vinyl decal stickers, skate branding, and high-contrast studio shadows.\nAct III: Decal sticker prompt code.',
        'slides': 6, 'trigger': 'DECAL', 'offer': 'Streetwear Decal Prompt Code'
    },
    {
        'day': 15, 'day_name': 'Monday', 'pillar': 'The 15-Second Retention Curve',
        'topic': 'How to Turn a 3-Second Hook into a Completed Watch & Comment',
        'hook': 'why 80% of viewers leave your carousel on slide 2.',
        'script': 'Act I: Hook dropoff analysis.\nAct II: The 15-second retention curve: 0-3s Hook -> 4-8s Tension -> 9-13s Reward -> 14-15s CTA.\nAct III: Retention curve checklist.',
        'slides': 8, 'trigger': 'HOOKS', 'offer': '25 Viral Retention Hook Scripts'
    },
    {
        'day': 16, 'day_name': 'Tuesday', 'pillar': 'Kallaway 6-Level Storytelling',
        'topic': 'From Level 1 Reporter to Level 6 Maestro: The Brand Storytelling Pyramid',
        'hook': 'where your brand sits on the 6 levels of storytelling determine your pricing power.',
        'script': 'Act I: Level 1 (Reporter) vs Level 2 (Illusionist CGI tricks).\nAct II: Deconstruct Level 5 (Translator) and Level 6 (Maestro Worldbuilding).\nAct III: The 6-Level Upgrade Roadmap.',
        'slides': 9, 'trigger': 'MAESTRO', 'offer': '6-Level Brand Storytelling Guide'
    },
    {
        'day': 17, 'day_name': 'Wednesday', 'pillar': 'Spring Physics & Elastic Bounce',
        'topic': 'The Damped Harmonic Spring Equation for Photoreal Video Motion',
        'hook': 'stiff keyframe animation is dead. mass, tension, and friction run 2026 cinema.',
        'script': 'Act I: Linear keyframe motion looks robotic.\nAct II: Damped harmonic spring math: mass 1.0, tension 180, friction 12.\nAct III: Spring math code snippets for Remotion & Web.',
        'slides': 7, 'trigger': 'SPRING', 'offer': 'Spring Physics Math Code Snippets'
    },
    {
        'day': 18, 'day_name': 'Thursday', 'pillar': 'High-Status Editorial Layouts',
        'topic': 'Graph Paper, Pastel Highlighter & Floating 3D Icons: Tactical Agency Grids',
        'hook': 'how top design agencies organize complex information into high-converting cards.',
        'script': 'Act I: Plain bullet lists fail to hold attention.\nAct II: Community graph paper, pastel highlighter blocks, and 3D icon anchors.\nAct III: Tactical Strategy Grid Figma Template.',
        'slides': 8, 'trigger': 'AGENCYGRID', 'offer': 'Tactical Strategy Grid Figma Template'
    },
    {
        'day': 19, 'day_name': 'Friday', 'pillar': 'Auteur Directing Case Study',
        'topic': 'Deconstructing Jonathan Glazer & Spike Jonze: Commercial Directing Secrets',
        'hook': 'what iconic film directors know about attention that advertisers miss.',
        'script': 'Act I: Advertising talks at people; auteur film pulls people into tension.\nAct II: Deconstructing visual pacing, sound drops, and atmospheric shadows.\nAct III: Auteur directing prompt swipe file.',
        'slides': 8, 'trigger': 'AUTEUR', 'offer': 'Auteur Directing Prompt Swipe File'
    },
    {
        'day': 20, 'day_name': 'Saturday', 'pillar': 'AI Commerce & Physical Moats',
        'topic': 'When Pixels Are Free, Story Is the Only Commercial Moat',
        'hook': 'in 2026, anyone can generate an 8k render. here is what actually sells physical goods.',
        'script': 'Act I: Commodity AI visual noise.\nAct II: The 3-part identity commerce engine.\nAct III: Physical product launching checklist.',
        'slides': 7, 'trigger': 'COMMERCE', 'offer': 'Physical Product Launch Playbook'
    },
    {
        'day': 21, 'day_name': 'Sunday', 'pillar': 'The 9-Year-Old Clarity Rule',
        'topic': 'Why Complex Visual Diagrams Win High-Ticket Buyers',
        'hook': 'if a 9-year-old cannot understand your slide in 2 seconds, your copy is broken.',
        'script': 'Act I: High-sounding jargon loses buyers.\nAct II: Translating dense concepts into 0.5px data cards, before/after splits, and 3-step flows.\nAct III: Visual clarity audit checklist.',
        'slides': 6, 'trigger': 'CLARITY', 'offer': 'Visual Clarity Audit Checklist'
    },
    {
        'day': 22, 'day_name': 'Monday', 'pillar': 'Y2K Direct Flash & Nightlife',
        'topic': 'Direct Flash 35mm Nightlife Photography: Raw Authenticity in Campaign Design',
        'hook': 'why harsh direct flash photography creates instant cult status.',
        'script': 'Act I: Soft studio lighting can feel generic and sterile.\nAct II: Direct flash highlights, deep vignetting, and Y2K nightlife energy.\nAct III: Direct flash prompt formula.',
        'slides': 8, 'trigger': 'DIRECTFLASH', 'offer': 'Direct Flash 35mm Prompt Code'
    },
    {
        'day': 23, 'day_name': 'Tuesday', 'pillar': 'Tactile Material Anatomy',
        'topic': 'Exploded Scent Pyramids & Viscosity Proof: Selling Physical Luxury Online',
        'hook': 'how to make a viewer feel the weight and scent of a product through a screen.',
        'script': 'Act I: E-commerce product pages are flat and unfeeling.\nAct II: Macro texture photography, ingredient origin teardowns, and swatch viscosity proof.\nAct III: Material Anatomy Carousel Template.',
        'slides': 8, 'trigger': 'MATERIAL', 'offer': 'Material Anatomy Carousel Template'
    },
    {
        'day': 24, 'day_name': 'Wednesday', 'pillar': 'The Double-Down Scaling Rule',
        'topic': 'How We Test 7 Hook Angles & Double Down on Outlier Winners',
        'hook': 'never guess what content will go viral. test 7 angles and double down.',
        'script': 'Act I: Relying on a single hook formula.\nAct II: The 7 rotating hook framework: Contrarian, Case Study, Blueprint, Comparison, Anatomy, Uncanny, Story Arc.\nAct III: Outlier tracking system.',
        'slides': 9, 'trigger': 'SCALING', 'offer': 'Outlier Content Tracking System'
    },
    {
        'day': 25, 'day_name': 'Thursday', 'pillar': '0.5x Fisheye Tech Commercials',
        'topic': 'Fisheye 0.5x Distorted Gadget Thrusts & Pop Color Badges',
        'hook': 'why 0.5x fisheye perspective creates hyper-energetic tech drops.',
        'script': 'Act I: Standard product angles lack energy.\nAct II: Extreme wide-angle lens distortion, yellow sale badges, and pop color accents.\nAct III: 0.5x Fisheye Tech Prompt Formula.',
        'slides': 6, 'trigger': 'FISHEYE', 'offer': '0.5x Fisheye Prompt Pack'
    },
    {
        'day': 26, 'day_name': 'Friday', 'pillar': 'High-Converting CTA Engineering',
        'topic': 'The Slide 8 DM Keyword Funnel: How to Get 500+ Qualified Leads Every Week',
        'hook': 'why your final slide CTA is failing to generate leads.',
        'script': 'Act I: Weak final slides.\nAct II: High-contrast 1-word DM trigger keywords, value offer clarity, and save/share calls.\nAct III: Slide 8 High-Converting CTA Matrix.',
        'slides': 8, 'trigger': 'CTAMATRIX', 'offer': 'Slide 8 CTA High-Converting Matrix'
    },
    {
        'day': 27, 'day_name': 'Saturday', 'pillar': 'The 3-Act Narrative Spine',
        'topic': 'The Act I Truth, Act II Tension, Act III Resolve Framework for Ads',
        'hook': 'the 3-act narrative spine that turns casual viewers into brand believers.',
        'script': 'Act I: Act I (The Unspoken Truth).\nAct II: Act II (Sensory Struggle & Float).\nAct III: Act III (The Physical Anchor).\nAct IV: Master 3-Act Ad Scripting Formula.',
        'slides': 9, 'trigger': 'THRERACT', 'offer': '3-Act Ad Scripting Blueprint'
    },
    {
        'day': 28, 'day_name': 'Sunday', 'pillar': 'Interactive Lead Magnet Architecture',
        'topic': 'How to Build 1-Click Copy Interactive HTML Playbooks for Your Audience',
        'hook': 'why sending static PDFs is outdated—and what interactive lead magnets do instead.',
        'script': 'Act I: Static PDF downloads get buried in downloads folders.\nAct II: Interactive HTML lead magnets with 1-click prompt copy buttons and live previews.\nAct III: Interactive HTML Template Kit.',
        'slides': 7, 'trigger': 'PLAYBOOK', 'offer': 'Interactive HTML Lead Magnet Kit'
    },
    {
        'day': 29, 'day_name': 'Monday', 'pillar': 'Auteur Visual Directing Retrospective',
        'topic': '30 Days of Visual Directing: What Worked, What Failed & What We Learned',
        'hook': 'we tested 30 visual directing strategies in 30 days. here are the 5 outlier winners.',
        'script': 'Act I: 30-day experiment overview.\nAct II: Top 3 winning visual styles & top 3 winning hook frameworks.\nAct III: The complete master visual directing vault.',
        'slides': 10, 'trigger': 'RETRO', 'offer': '30-Day Master Visual Directing Vault'
    },
    {
        'day': 30, 'day_name': 'Tuesday', 'pillar': 'The Defensible Commercial Moat',
        'topic': 'Building a 7-Figure Luxury Brand World: The Complete 2026 Master Plan',
        'hook': 'the exact blueprint to build an untouchable luxury visual moat in 2026.',
        'script': 'Act I: The future of luxury brand worldbuilding.\nAct II: Combining Auteur Cinema + AI Visual Prompt Cookbook + OpenReply 24/7 Funnels.\nAct III: Master 2026 Brand Worldbuilding Blueprint.',
        'slides': 10, 'trigger': 'CIEL', 'offer': 'Master 2026 Brand Worldbuilding Blueprint'
    }
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '30-Day Content Calendar'
ws.views.sheetView[0].showGridLines = True

headers = [
    'Day #', 'Day of Week', 'Content Pillar', 'Topic Title',
    'Headline Hook (Slide 1)', 'Deconstructed Script & Story Arc',
    'Cookbook Style Slug', 'Cookbook Style Name',
    'Slide Count', 'DM Trigger Keyword', 'Lead Magnet Offer', 'Destination Link'
]

header_fill = PatternFill(start_color='0A0A0C', end_color='0A0A0C', fill_type='solid')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
regular_font = Font(name='Calibri', size=10, color='111111')
bold_font = Font(name='Calibri', size=10, bold=True, color='111111')
code_font = Font(name='Consolas', size=9, bold=True, color='2B6CB0')

thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

ws.append(headers)
for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for idx, d in enumerate(days_data):
    style_idx = idx % len(style_slugs)
    slug = style_slugs[style_idx]
    style_name = slug.replace('-', ' ').title()
    
    row = [
        d['day'],
        d['day_name'],
        d['pillar'],
        d['topic'],
        d['hook'],
        d['script'],
        slug,
        style_name,
        d['slides'],
        d['trigger'],
        d['offer'],
        'https://t.me/projectciel'
    ]
    
    ws.append(row)
    row_num = idx + 2
    is_alt = (idx % 2 == 1)
    
    for col_num in range(1, len(row) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.border = thin_border
        if is_alt:
            cell.fill = alt_fill
            
        if col_num in [1, 9]:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_num in [2, 10]:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_num in [7]:
            cell.font = code_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
        elif col_num in [12]:
            cell.font = Font(name='Calibri', size=10, color='3182CE', underline='single')
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.font = regular_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

col_widths = {
    1: 8, 2: 12, 3: 26, 4: 38, 5: 36, 6: 50,
    7: 35, 8: 30, 9: 12, 10: 16, 11: 35, 12: 24
}

for col_num, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

ws.row_dimensions[1].height = 28
for r in range(2, len(days_data) + 2):
    ws.row_dimensions[r].height = 45

out_dir = r'out'
os.makedirs(out_dir, exist_ok=True)
excel_path = os.path.join(out_dir, '30_DAY_CAROUSEL_CONTENT_CALENDAR.xlsx')
wb.save(excel_path)
print(f'[SUCCESS] Excel spreadsheet saved to: {excel_path}')

# Save JSON version
json_path = os.path.join(out_dir, '30_DAY_CAROUSEL_CONTENT_CALENDAR.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(days_data, f, indent=2)
print(f'[SUCCESS] JSON data saved to: {json_path}')
